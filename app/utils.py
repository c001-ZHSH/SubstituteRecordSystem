import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def num2zh(num):
    """Simple converter for amount to Chinese uppercase string."""
    num_str = str(int(num))
    zh_num = {'0': '零', '1': '壹', '2': '貳', '3': '參', '4': '肆', 
              '5': '伍', '6': '陸', '7': '柒', '8': '捌', '9': '玖'}
    zh_unit = ['', '拾', '佰', '仟', '萬', '拾', '佰', '仟', '億']
    
    res = ""
    zero_flag = False
    
    for i, d in enumerate(num_str):
        unit_idx = len(num_str) - 1 - i
        if d == '0':
            zero_flag = True
            # For 萬 or 億 we still append the unit even with zeros if previous segment wasn't all zero
            if unit_idx in [4, 8]:
                 res += zh_unit[unit_idx]
        else:
            if zero_flag:
                res += zh_num['0']
                zero_flag = False
            res += zh_num[d] + zh_unit[unit_idx]
            
    if not res:
        res = "零"
    # Basic strip for trailing zeros and units logic
    res = res.replace('零萬', '萬').replace('零億', '億')
    while res.endswith('零'):
        res = res[:-1]
    return f"新臺幣 {res}元整"

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_font(ws, cell_range, font_name="微軟正黑體", size=11, bold=False):
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(name=font_name, size=size, bold=bold)

def set_alignment(ws, cell_range, h="center", v="center", wrap_text=False):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap_text)

def generate_substitute_list_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "代課清單"

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "代課清單 (系統產出)"
    ws['A1'].font = Font(name="微軟正黑體", size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["資料編號", "請假教師姓名", "假別", "代課日期", "代課教師姓名", "節次", "科目", "班級", "節數", "備註"]
    ws.append(headers)

    # Group records by LeaveRecord.id
    from collections import defaultdict
    groups = defaultdict(list)
    for r in records:
        # Completely exclude swapped records from the substitute list export
        if r.is_swapped in [True, 1, '1', 'true', 'True']:
            continue
        groups[r.leave_record_id].append(r)

    start_row = 3
    idx = 1
    
    for lid, group_records in groups.items():
        master = group_records[0].leave_record
        row_count = len(group_records)
        end_row = start_row + row_count - 1

        # Write rows
        for i, sub in enumerate(group_records):
            current_row = start_row + i
            if i == 0:
                ws.cell(row=current_row, column=1, value=idx)
                ws.cell(row=current_row, column=2, value=master.teacher_name)
                
                leave_info = master.leave_reason
                if master.leave_reason == "公假" and master.approval_number:
                    leave_info += f" ( {master.approval_number} )"
                elif master.approval_number:
                    leave_info += f" ({master.approval_number})"
                    
                ws.cell(row=current_row, column=3, value=leave_info)
            
            ws.cell(row=current_row, column=4, value=sub.substitute_date)
            ws.cell(row=current_row, column=5, value=sub.substitute_teacher)
            ws.cell(row=current_row, column=6, value=sub.periods)
            ws.cell(row=current_row, column=7, value=sub.subject)
            ws.cell(row=current_row, column=8, value=sub.class_name)
            ws.cell(row=current_row, column=9, value=sub.period_count)
            
            remark = sub.remarks or ''
            # Handle possible python types from SQLAlchemy Boolean column
            is_moe = sub.is_moe_subsidized in [True, 1, '1', 'true', 'True']
            is_swapped = sub.is_swapped in [True, 1, '1', 'true', 'True']
            
            if is_swapped:
                remark = f"{remark} (調課)".strip()
            elif is_moe:
                remark = f"{remark} (教育部補助)".strip()
                
            ws.cell(row=current_row, column=10, value=remark)

        # Merge cells for master data
        if row_count > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        start_row += row_count
        idx += 1

    # End
    end_data_row = start_row - 1
    
    # Signature row
    sig_row = start_row
    
    # 承辦人 (Column A) and its empty stamp box (Column B)
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row+2, end_column=1)
    ws.cell(row=sig_row, column=1, value="承辦人")
    ws.merge_cells(start_row=sig_row, start_column=2, end_row=sig_row+2, end_column=2)
    
    # 輔導主任
    ws.merge_cells(start_row=sig_row, start_column=3, end_row=sig_row, end_column=4)
    ws.cell(row=sig_row, column=3, value="輔導主任")
    ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=7)
    
    # 教務主任
    ws.merge_cells(start_row=sig_row+1, start_column=3, end_row=sig_row+1, end_column=4)
    ws.cell(row=sig_row+1, column=3, value="教務主任")
    ws.merge_cells(start_row=sig_row+1, start_column=5, end_row=sig_row+1, end_column=7)
    
    # 人事主任
    ws.merge_cells(start_row=sig_row+2, start_column=3, end_row=sig_row+2, end_column=4)
    ws.cell(row=sig_row+2, column=3, value="人事主任")
    ws.merge_cells(start_row=sig_row+2, start_column=5, end_row=sig_row+2, end_column=7)

    # 校長 (Principal)
    ws.merge_cells(start_row=sig_row, start_column=8, end_row=sig_row+2, end_column=8)
    ws.cell(row=sig_row, column=8, value="校長")
    ws.cell(row=sig_row, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=sig_row, start_column=9, end_row=sig_row+2, end_column=10)
    
    # Adjust row heights to provide room for stamps
    for r in range(sig_row, sig_row+3):
        ws.row_dimensions[r].height = 40
    
    # Formatting
    set_border(ws, f'A1:J{sig_row+2}')
    set_font(ws, f'A1:J{sig_row+2}')
    set_alignment(ws, f'A1:J{sig_row+2}')

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 15

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path

def _build_payment_sheet(ws, records, unit_price, teacher_deductions, title="代課教師超鐘點費印領清冊"):
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = title
    ws['A1'].font = Font(name="微軟正黑體", size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Header Row 1
    ws.row_dimensions[2].height = 25
    ws.merge_cells('A2:A3')
    ws['A2'] = "序號"
    ws.merge_cells('B2:B3')
    ws['B2'] = "教師姓名"
    ws.merge_cells('C2:C3')
    ws['C2'] = "上課時間"
    
    ws.merge_cells('D2:F2')
    ws['D2'] = "鐘點費"
    ws.merge_cells('G2:H2')
    ws['G2'] = "代扣款(元)"
    
    ws.merge_cells('I2:I3')
    ws['I2'] = "實領金額(元)"

    # Header Row 2
    ws['D3'] = "節數"
    ws['E3'] = "單價(元)"
    ws['F3'] = "合計(元)"
    ws['G3'] = "健保"
    ws['H3'] = "勞保"

    total_periods = 0
    total_amount = 0
    total_health_ins = 0
    total_labor_ins = 0

    applied_deductions = set() # Track teachers that already got deduction applied

    # Write Data
    start_row = 4
    for i, sub in enumerate(records):
        row = start_row + i
        time_str = f"{sub.substitute_date} 第{sub.periods}節 {sub.subject}"
        teacher_name = sub.substitute_teacher
        
        periods = int(sub.period_count)
        fee_subtotal = periods * unit_price
        
        # Apply deductions only once per teacher in the receipt
        health_ins = 0
        labor_ins = 0
        if teacher_name not in applied_deductions:
            teacher_data = teacher_deductions.get(teacher_name, {})
            health_ins = teacher_data.get('health', 0)
            labor_ins = teacher_data.get('labor', 0)
            applied_deductions.add(teacher_name)
            
        total_health_ins += health_ins
        total_labor_ins += labor_ins

        actual_total = fee_subtotal - health_ins - labor_ins

        total_periods += periods
        total_amount += actual_total

        ws.cell(row=row, column=1, value=i+1)
        ws.cell(row=row, column=2, value=teacher_name)
        ws.cell(row=row, column=3, value=time_str)
        ws.cell(row=row, column=4, value=periods)
        ws.cell(row=row, column=5, value=unit_price)
        ws.cell(row=row, column=6, value=fee_subtotal)
        ws.cell(row=row, column=7, value=health_ins)
        ws.cell(row=row, column=8, value=labor_ins)
        ws.cell(row=row, column=9, value=actual_total)

    # Empty rows per screenshot
    current_row = start_row + len(records)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.cell(row=current_row, column=1, value="以下空白")
    current_row += 1

    # Total Sum
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.cell(row=current_row, column=1, value="合計")
    ws.cell(row=current_row, column=4, value=total_periods)
    ws.cell(row=current_row, column=5, value=unit_price)
    ws.cell(row=current_row, column=6, value=total_periods * unit_price)
    ws.cell(row=current_row, column=7, value=total_health_ins)
    ws.cell(row=current_row, column=8, value=total_labor_ins)
    ws.cell(row=current_row, column=9, value=total_amount)
    
    current_row += 1

    # Number to Word
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    zh_amount_str = num2zh(total_amount)
    ws.cell(row=current_row, column=1, value=f"總計 {total_periods} 節，${total_amount} 元整，{zh_amount_str}")

    # Formatting
    set_border(ws, f'A1:I{current_row}')
    set_font(ws, f'A1:I{current_row}')
    set_alignment(ws, f'A1:I{current_row}')
    
    # Signatures placeholder
    current_row += 2
    ws.cell(row=current_row, column=1, value="製表")
    ws.cell(row=current_row, column=3, value="教務處")
    ws.cell(row=current_row, column=6, value="校長")
    
    current_row += 2
    ws.cell(row=current_row, column=1, value="單位主管")
    ws.cell(row=current_row, column=3, value="總務處")
    
    current_row += 2
    ws.cell(row=current_row, column=3, value="人事室")
    
    current_row += 2
    ws.cell(row=current_row, column=3, value="會計室")

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15

def generate_payment_excel(records, unit_price, teacher_deductions):
    wb = Workbook()
    
    regular_records = []
    moe_records = []
    
    for r in records:
        # Exclude Swapped Classes from payroll altogether 
        if r.is_swapped in [True, 1, '1', 'true', 'True']:
            continue
            
        if r.is_moe_subsidized in [True, 1, '1', 'true', 'True']:
            moe_records.append(r)
        else:
            regular_records.append(r)
    
    # Generate Regular Sheet
    ws_regular = wb.active
    ws_regular.title = "印領清冊"
    if regular_records:
        _build_payment_sheet(ws_regular, regular_records, unit_price, teacher_deductions, "代課教師超鐘點費印領清冊")
    else:
        ws_regular['A1'] = "無一般代課紀錄"

    # Generate MOE Sheet if necessary
    if moe_records:
        ws_moe = wb.create_sheet(title="印領清冊(教育部補助)")
        _build_payment_sheet(ws_moe, moe_records, unit_price, teacher_deductions, "代課教師超鐘點費印領清冊(教育部補助)")

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path
